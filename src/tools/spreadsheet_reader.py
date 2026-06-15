"""Load and filter provider license data from an Excel workbook."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def load_spreadsheet(
    path: str | Path,
    target_month: str,
) -> tuple[list[dict], dict[str, dict]]:
    """
    Returns (licenses, providers) filtered to licenses expiring in target_month.

    licenses: list of dicts — initials, state, license_type, expiration_date (date)
    providers: dict keyed by initials — full_name, credential, email, np_compact (bool)
    """
    target_year, target_month_num = _parse_target_month(target_month)

    wb = openpyxl.load_workbook(path, data_only=True)

    ws_prov = wb["providers"]
    prov_headers = [c.value for c in next(ws_prov.iter_rows(min_row=1, max_row=1))]
    providers: dict[str, dict] = {}
    for row in ws_prov.iter_rows(min_row=2, values_only=True):
        r = dict(zip(prov_headers, row))
        if r.get("initials"):
            providers[r["initials"]] = {
                "full_name": r.get("full_name", ""),
                "credential": r.get("credential", ""),
                "email": r.get("email", ""),
                "np_compact": str(r.get("np_compact", "No")).strip().lower() == "yes",
            }

    ws_lic = wb["licenses"]
    lic_headers = [c.value for c in next(ws_lic.iter_rows(min_row=1, max_row=1))]
    licenses: list[dict] = []
    for row in ws_lic.iter_rows(min_row=2, values_only=True):
        r = dict(zip(lic_headers, row))
        exp = _parse_date(r.get("expiration_date"))
        if exp is None:
            continue
        if exp.year == target_year and exp.month == target_month_num:
            licenses.append(
                {
                    "initials": r["initials"],
                    "state": r["state"],
                    "license_type": r["license_type"],
                    "expiration_date": exp,
                }
            )

    return licenses, providers


def _parse_target_month(target_month: str) -> tuple[int, int]:
    s = target_month.strip()
    for fmt in ("%B %Y", "%b %Y", "%m/%Y", "%Y-%m"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.year, dt.month
        except ValueError:
            continue
    raise ValueError(f"Cannot parse target month: {target_month!r}")
